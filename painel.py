import requests
import base64
import os
import json
import socket
from flask import Flask, request, jsonify, render_template_string
from datetime import datetime, timedelta

app = Flask(__name__)

API_BASE_URL = "https://consultadanfe.com/api/v1/consulta"
CACHE_DIR = "cache_notas"
if not os.path.exists(CACHE_DIR):
    os.makedirs(CACHE_DIR)

# --- NOVO LAYOUT DO TEMPLATE MÃE (DIVISOR MODERNO COM GLASSMORPHISM) ---
HTML_TEMPLATE = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <title>Painel Integrado Executivo</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body, html { width: 100%; height: 100%; overflow: hidden; background: #0f172a; font-family: 'Segoe UI', system-ui, sans-serif; }
        
        .dashboard-container { 
            display: flex; 
            width: 100%; 
            height: 100%; 
            background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
            padding: 10px;
            gap: 12px;
        }
        
        /* Divisão idêntica 50/50 com efeito de cards flutuantes independentes */
        .lado-esquerdo, .lado-direito { 
            flex: 1; 
            height: 100%; 
            border: none; 
            border-radius: 12px;
            box-shadow: 0 20px 25px -5px rgba(0, 0, 0, 0.3), 0 10px 10px -5px rgba(0, 0, 0, 0.2);
            background: #ffffff;
            transition: all 0.3s ease;
        }

        .lado-esquerdo:hover, .lado-direito:hover {
            box-shadow: 0 25px 30px -5px rgba(0, 0, 0, 0.4);
        }
    </style>
</head>
<body>

    <div class="dashboard-container">
        <iframe src="/api/cnpj_html" class="lado-esquerdo"></iframe>
        
        <iframe src="/index_notas" class="lado-direito"></iframe>
    </div>

</body>
</html>
"""

# --- NOVO LAYOUT DO PROJETO DE NOTAS (COMPLETAMENTE REESTILIZADO) E COM O MOTOR PARALELO CORRIGIDO ---
HTML_NOTAS_ORIGINAL = """
<!DOCTYPE html>
<html lang="pt-BR">
<head>
    <meta charset="UTF-8">
    <title>Fast Loader & Conciliador Exact</title>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/jszip/3.10.1/jszip.min.js"></script>
    <style>
        :root { 
            --primary: #3b82f6; 
            --primary-hover: #2563eb;
            --bg: #f8fafc; 
            --success: #10b981; 
            --success-bg: #ecfdf5;
            --danger: #ef4444;
            --danger-bg: #fef2f2;
            --pdf: #f43f5e; 
            --xml: #0284c7; 
            --cache: #8b5cf6; 
            --orange: #f97316; 
            --zip: #0d9488;
            --slate-dark: #0f172a;
            --slate-text: #475569;
            --border: #e2e8f0;
        }
        
        body { 
            font-family: 'Segoe UI', -apple-system, BlinkMacSystemFont, Roboto, sans-serif; 
            background: var(--bg); 
            padding: 20px; 
            color: var(--slate-text); 
            margin: 0;
            overflow-y: auto;
        }
        
        .container { 
            max-width: 100%; 
            margin: 0 auto; 
            background: white; 
            padding: 24px; 
            border-radius: 12px; 
            box-shadow: 0 1px 3px rgba(0,0,0,0.05); 
        }
        
        /* Abas Estilo Corporativo Moderno */
        .tabs { 
            display: flex; 
            gap: 6px; 
            margin-bottom: 24px; 
            border-bottom: 1px solid var(--border); 
            padding-bottom: 8px; 
        }
        .tab-btn { 
            padding: 10px 18px; 
            font-weight: 600; 
            border: none; 
            background: none; 
            cursor: pointer; 
            color: #64748b; 
            border-radius: 6px; 
            transition: all 0.2s ease; 
            font-size: 13.5px; 
        }
        .tab-btn:hover { 
            background: #f1f5f9; 
            color: var(--slate-dark); 
        }
        .tab-btn.active { 
            background: #eff6ff; 
            color: var(--primary); 
        }
        
        .tab-content { display: none; animation: fadeIn 0.3s ease; }
        .tab-content.active { display: block; }
        @keyframes fadeIn { from { opacity: 0; transform: translateY(4px); } to { opacity: 1; transform: translateY(0); } }

        h1 { margin-top: 0; color: var(--slate-dark); font-size: 20px; font-weight: 700; text-align: center; margin-bottom: 4px; letter-spacing: -0.5px; }
        p.subtitle { text-align: center; color: #64748b; margin-bottom: 24px; font-size: 13px; }
        
        textarea { 
            width: 100%; 
            border: 1px solid #cbd5e1; 
            border-radius: 8px; 
            padding: 14px; 
            font-family: 'SFMono-Regular', Consolas, monospace; 
            box-sizing: border-box; 
            font-size: 13px; 
            margin-top: 10px; 
            background: #fafafa; 
            transition: all 0.2s; 
            resize: vertical;
        }
        textarea:focus { border-color: var(--primary); outline: none; background: white; box-shadow: 0 0 0 3px rgba(59, 130, 246, 0.15); }
        
        .btn-principal { 
            border: none; 
            padding: 13px; 
            border-radius: 8px; 
            cursor: pointer; 
            font-weight: 600; 
            color: white; 
            background: var(--primary); 
            width: 100%; 
            font-size: 14px; 
            margin-top: 15px; 
            transition: background 0.2s; 
            box-shadow: 0 1px 2px rgba(0,0,0,0.05);
        }
        .btn-principal:hover { background: var(--primary-hover); }
        
        .lote-actions-grid { display: flex; gap: 10px; margin-top: 15px; }
        .btn-lote { 
            flex: 1; 
            border: none; 
            padding: 11px; 
            border-radius: 8px; 
            cursor: pointer; 
            font-weight: 600; 
            color: white; 
            font-size: 13px; 
            transition: filter 0.2s; 
            text-align: center; 
            box-shadow: 0 1px 2px rgba(0,0,0,0.05);
        }
        .btn-lote:hover { filter: brightness(0.92); }
        
        /* Upload Zones Limpas */
        .upload-grid { display: flex; gap: 14px; margin: 20px 0; }
        .file-box { 
            border: 2px dashed #e2e8f0; 
            padding: 20px; 
            text-align: center; 
            border-radius: 8px; 
            background: #f8fafc; 
            flex: 1; 
            cursor: pointer; 
            transition: all 0.2s ease; 
        }
        .file-box:hover { border-color: var(--primary); background: #f0f7ff; }
        .file-box span { font-size: 13px; font-weight: 600; color: #475569; }
        
        /* Dashboard de Resumo Compacto */
        .summary-dashboard { display: flex; gap: 12px; margin-bottom: 20px; }
        .summary-card { background: #f8fafc; border: 1px solid var(--border); padding: 12px 16px; border-radius: 8px; flex: 1; }
        .summary-card .title { font-size: 10px; text-transform: uppercase; color: #64748b; font-weight: 700; letter-spacing: 0.5px; }
        .summary-card .value { font-size: 18px; font-weight: 700; color: var(--slate-dark); margin-top: 2px; }

        /* Tabelas Polidas */
        .table-container { border: 1px solid var(--border); border-radius: 8px; overflow: hidden; margin-top: 20px; }
        table { width: 100%; border-collapse: collapse; font-size: 13px; table-layout: fixed; background: white; }
        th, td { padding: 12px 14px; text-align: left; vertical-align: middle; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
        th { background: #f8fafc; color: #64748b; font-weight: 600; text-transform: uppercase; font-size: 10px; letter-spacing: 0.5px; border-bottom: 1px solid var(--border); }
        
        tbody tr { border-bottom: 1px solid #f1f5f9; transition: background 0.15s; }
        tbody tr:hover { background: #f8fafc; }
        tbody tr:last-child { border-bottom: none; }
        
        .badge { display: inline-block; padding: 4px 8px; border-radius: 4px; font-weight: 700; font-size: 11px; text-align: center; }
        .badge-prp { background: #fff7ed; color: var(--orange); border: 1px solid #ffedd5; cursor: pointer; }
        .badge-prp:hover { border-color: var(--primary); color: var(--primary); }
        .badge-success { background: var(--success-bg); color: var(--success); }
        .badge-danger { background: var(--danger-bg); color: var(--danger); font-weight: 600; }
        
        .btn-mini { padding: 5px 10px; font-size: 11px; border-radius: 4px; border: none; color: white; cursor: pointer; font-weight: 700; margin-right: 4px; transition: opacity 0.2s; }
        .btn-mini:hover { opacity: 0.9; }
        .btn-pdf { background: var(--pdf); }
        .btn-xml { background: var(--xml); }
        
        .copy-clickable { font-family: 'SFMono-Regular', Consolas, monospace; font-size: 11.5px; color: #475569; cursor: pointer; }
        .copy-clickable:hover { color: var(--primary); text-decoration: underline; }
        
        .copy-toast { position: fixed; background: #0f172a; color: white; padding: 6px 12px; font-size: 11px; font-weight: 600; border-radius: 4px; pointer-events: none; z-index: 9999; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.1); }
        #log, #log-conciliador { text-align: center; font-weight: 600; margin-top: 15px; color: var(--primary); font-size: 13px; }
    </style>
</head>
<body>
    <div class="container">
        <div class="tabs">
            <button class="tab-btn active" onclick="switchTab('busca')">🔍 1º Passo: Buscar Notas</button>
            <button class="tab-btn" onclick="switchTab('conciliador')">📊 2º Passo: Identificar Comprador & PRP</button>
        </div>

        <div id="tab-busca" class="tab-content active">
            <h1>Consulta Cadenciada com Cache</h1>
            <p class="subtitle">Insira as chaves de acesso para download automático e alimentação da memória local.</p>
            <textarea id="chaves" rows="5" placeholder="Cole as chaves de acesso aqui (uma por linha)..."></textarea>
            <button class="btn-principal" onclick="consultarNotasSequencial()">Iniciar Consulta Lote</button>
            
            <div class="lote-actions-grid" id="acoes-lote-container" style="display: none;">
                <button class="btn-lote" style="background: var(--pdf);" onclick="abrirTodosPdfsLote()">🚨 Abrir Todos os PDFs do Lote</button>
                <button class="btn-lote" style="background: var(--zip);" onclick="baixarTodosXmlsZip()">📦 Baixar Todos os XMLs (.ZIP)</button>
            </div>

            <div id="log"></div>
            <div class="table-container" id="container-tabela-busca" style="display:none">
                <table id="tabela">
                    <thead>
                        <tr><th style="width: 8%;">#</th><th style="width: 52%;">Chave</th><th style="width: 20%;">Status</th><th style="width: 20%;">Ações</th></tr>
                    </thead>
                    <tbody id="corpo"></tbody>
                </table>
            </div>
        </div>

        <div id="tab-conciliador" class="tab-content">
            <h1>Vinculador de Responsáveis e PRP</h1>
            <p class="subtitle">Análise inteligível baseada na data de emissão do XML cruzada com portadores e períodos de viagem.</p>
            
            <div class="upload-grid">
                <div class="file-box" onclick="document.getElementById('excel_extrato').click()">
                    <span id="label-extrato">📄 1. Planilha Conta Simples (.xlsx)</span>
                    <input type="file" id="excel_extrato" accept=".xlsx" style="display:none" onchange="atualizarLabel('extrato')">
                </div>
                
                <div class="file-box" onclick="document.getElementById('excel_controle').click()">
                    <span id="label-controle">📄 2. Planilha Pedidos de Viagem (.xlsx)</span>
                    <input type="file" id="excel_controle" accept=".xlsx" style="display:none" onchange="atualizarLabel('controle')">
                </div>
            </div>

            <button class="btn-principal" style="background:var(--orange);" onclick="executarDuplaConciliacao()">🚀 Mapear Compradores e PRPs do Lote</button>
            <div id="log-conciliador"></div>

            <div class="summary-dashboard" id="dashboard-conciliacao" style="display:none; margin-top:20px;">
                <div class="summary-card">
                    <div class="title">Notas Processadas</div>
                    <div class="value" id="dash-total-notas">0</div>
                </div>
                <div class="summary-card">
                    <div class="title">Status do Motor</div>
                    <div class="value" style="color:var(--success);">Ativo e Cruzado</div>
                </div>
            </div>

            <div class="table-container" id="container-tabela-conciliacao" style="display:none">
                <table id="tabela-conciliacao">
                    <thead>
                        <tr>
                            <th style="width: 12%; text-align: center;">Data (Nota)</th>
                            <th style="width: 11%; text-align: center;">Nº da Nota</th>
                            <th style="width: 26%;">🔑 Chave de Acesso</th>
                            <th style="width: 13%;">Valor NF</th>
                            <th style="width: 23%;">👤 Quem Comprou (Conta Simples)</th>
                            <th style="width: 15%;">🏢 PRP Vinculada</th>
                        </tr>
                    </thead>
                    <tbody id="corpo-conciliacao"></tbody>
                </table>
            </div>
        </div>
    </div>

    <script>
        function switchTab(tabId) {
            document.querySelectorAll('.tab-btn').forEach(b => b.classList.remove('active'));
            document.querySelectorAll('.tab-content').forEach(c => c.classList.remove('active'));
            if(event) event.target.classList.add('active');
            document.getElementById('tab-' + tabId).classList.add('active');
        }
        
        function atualizarLabel(tipo) {
            const input = document.getElementById('excel_' + tipo);
            if(input.files.length) { document.getElementById('label-' + tipo).innerText = `✅ ${input.files[0].name}`; }
        }
        
        function copiarTextoRapido(elemento, evento) {
            let texto = elemento.innerText.trim();
            if(!texto || texto === '-' || texto.includes("Não Localizado")) return;
            navigator.clipboard.writeText(texto).then(() => {
                const toast = document.createElement('div');
                toast.className = 'copy-toast'; toast.innerText = '📋 Copiado!';
                toast.style.left = (evento.clientX + 10) + 'px'; toast.style.top = (evento.clientY - 10) + 'px';
                document.body.appendChild(toast);
                setTimeout(() => toast.remove(), 1000);
            });
        }
        
        const sleep = ms => new Promise(r => setTimeout(r, ms));
        let cacheNotas = {}; let chavesDoLoteAtual = [];
        
        function gerarBlobPdf(base64Data) {
            const byteCharacters = atob(base64Data); const byteNumbers = new Array(byteCharacters.length);
            for (let i = 0; i < byteCharacters.length; i++) byteNumbers[i] = byteCharacters.charCodeAt(i);
            return new Blob([new Uint8Array(byteNumbers)], { type: 'application/pdf' });
        }
        
        function abrirPdf(chave) { if (cacheNotas[chave] && cacheNotas[chave].pdf) { window.open(URL.createObjectURL(gerarBlobPdf(cacheNotas[chave].pdf))); } }
        
        function baixarXml(chave) {
            if (cacheNotas[chave] && cacheNotas[chave].xml) {
                const link = document.createElement('a'); link.href = URL.createObjectURL(new Blob([cacheNotas[chave].xml], { type: 'text/xml' }));
                link.download = `${chave}.xml`; link.click();
            }
        }
        
        function abrirTodosPdfsLote() {
            chavesDoLoteAtual.forEach(chave => { if(cacheNotas[chave] && cacheNotas[chave].pdf) { window.open(URL.createObjectURL(gerarBlobPdf(cacheNotas[chave].pdf))); } });
        }
        
        function baixarTodosXmlsZip() {
            const zip = new JSZip(); let count = 0;
            chavesDoLoteAtual.forEach(chave => { if(cacheNotas[chave] && cacheNotas[chave].xml) { zip.file(`${chave}.xml`, cacheNotas[chave].xml); count++; } });
            if(count === 0) return alert("Nenhum XML localizado.");
            zip.generateAsync({ type: 'blob' }).then(content => {
                const link = document.createElement('a'); link.href = URL.createObjectURL(content);
                link.download = `LOTE_XML_${new Date().toISOString().slice(0,10)}.zip`; link.click();
            });
        }

        async function consultarNotasSequencial() {
            // Utilizando \\r?\\n para o Python não interpretar e manter a regex no JS
            const chaves = document.getElementById('chaves').value.split(/\\r?\\n/).map(c => c.trim()).filter(c => c.length == 44);
            
            if (!chaves.length) return alert("Insira chaves válidas.");
            
            chavesDoLoteAtual = chaves;
            document.getElementById('acoes-lote-container').style.display = 'none';
            document.getElementById('container-tabela-busca').style.display = 'block';
            
            const corpo = document.getElementById('corpo'); 
            corpo.innerHTML = '';
            
            for (let i = 0; i < chaves.length; i++) {
                corpo.innerHTML += `<tr><td>${i+1}</td><td class="copy-clickable">${chaves[i]}</td><td id="st-${i}">⏳ Na fila...</td><td id="ac-${i}">-</td></tr>`;
            }

            document.getElementById('log').innerText = "🚀 Processando lote em paralelo...";

            const limiteConcorrencia = 5; 
            let indiceGlobal = 0;

            async function dispararWorker() {
                while (indiceGlobal < chaves.length) {
                    const i = indiceGlobal; 
                    const bundle = chaves[i];
                    indiceGlobal++; 

                    try {
                        document.getElementById(`st-${i}`).innerHTML = `🔄 Consultando...`;
                        
                        const r = await fetch('/api/consulta_direta', { 
                            method: 'POST', 
                            headers: {'Content-Type': 'application/json'}, 
                            body: JSON.stringify({ chave: bundle }) 
                        });
                        const d = await r.json();
                        
                        if (d.success) {
                            cacheNotas[bundle] = { xml: d.xml, pdf: d.pdf };
                            document.getElementById(`st-${i}`).innerHTML = d.from_cache ? `<span class="badge" style="background:#f3e8ff; color:var(--cache);">⚡ Cache</span>` : `<span class="badge badge-success">✅ Pronto</span>`;
                            document.getElementById(`ac-${i}`).innerHTML = `<button class="btn-mini btn-pdf" onclick="abrirPdf('${bundle}')">PDF</button> <button class="btn-mini btn-xml" onclick="baixarXml('${bundle}')">XML</button>`;
                        } else { 
                            document.getElementById(`st-${i}`).innerHTML = `<span class="badge badge-danger">❌ ${d.reason}</span>`; 
                        }
                    } catch { 
                        document.getElementById(`st-${i}`).innerHTML = `<span class="badge badge-danger">❌ Erro</span>`; 
                    }
                }
            }

            const workers = [];
            for (let w = 0; w < limiteConcorrencia; w++) {
                workers.push(dispararWorker());
            }

            await Promise.all(workers);

            document.getElementById('log').innerText = "✅ Lote Concluído com Sucesso!";
            document.getElementById('acoes-lote-container').style.display = 'flex';
        }

        async function executarDuplaConciliacao() {
            const fileExtrato = document.getElementById('excel_extrato'); const fileControle = document.getElementById('excel_controle');
            if (!chavesDoLoteAtual.length) return alert("Faça primeiro a busca de notas na Aba 1.");
            if (!fileExtrato.files.length || !fileControle.files.length) return alert("Selecione ambas as planilhas.");
            
            document.getElementById('log-conciliador').innerText = "🤖 Processando cruzamentos...";
            const formData = new FormData();
            formData.append('extrato', fileExtrato.files[0]); formData.append('controle', fileControle.files[0]);
            formData.append('chaves_lote', JSON.stringify(chavesDoLoteAtual));
            
            try {
                const response = await fetch('/api/conciliar_duplo_restrito', { method: 'POST', body: formData });
                const result = await response.json();
                if (!result.success) { document.getElementById('log-conciliador').innerText = `❌ ${result.reason}`; return; }
                
                document.getElementById('dash-total-notas').innerText = result.results.length;
                document.getElementById('dashboard-conciliacao').style.display = 'flex';
                document.getElementById('container-tabela-conciliacao').style.display = 'block';
                
                const corpo = document.getElementById('corpo-conciliacao'); corpo.innerHTML = '';
                result.results.forEach(item => {
                    let prpBadge = `<span class="badge badge-prp" title="Clique para copiar a PRP" onclick="copiarTextoRapido(this, event)">%${item.prp}</span>`;
                    if(item.prp.includes("Não localizado")) prpBadge = `<span class="badge badge-danger">Não Localizado</span>`;
                    let compradorStyle = item.comprador.includes("Não localizado") ? 'color:#94a3b8; font-weight:normal;' : 'color:var(--slate-dark); font-weight:600;';
                    corpo.innerHTML += `<tr><td style="text-align: center; font-weight:600; color:#475569;" class="copy-clickable" onclick="copiarTextoRapido(this, event)">${item.data_nota_pt || '-'}</td><td style="text-align: center;" class="copy-clickable" onclick="copiarTextoRapido(this, event)">${item.numero_nota || '-'}</td><td class="copy-clickable" style="color:#64748b;" onclick="copiarTextoRapido(this, event)">${item.chave}</td><td style="font-weight:700; color:var(--primary);">R$ ${item.valor_nota.toFixed(2)}</td><td style="${compradorStyle}">${item.comprador}</td><td>${prpBadge}</td></tr>`;
                });
                document.getElementById('log-conciliador').innerText = "✅ Mapeamento finalizado!";
            } catch { document.getElementById('log-conciliador').innerText = "❌ Erro ao cruzar dados."; }
        }
    </script>
</body>
</html>
"""

# --- ROTAS DE EXIBIÇÃO ---

@app.route('/')
def index():
    return render_template_string(HTML_TEMPLATE)

@app.route('/index_notas')
def index_notas():
    return render_template_string(HTML_NOTAS_ORIGINAL)

@app.route('/api/cnpj_html')
def api_cnpj_html():
    try:
        with open('CONSULTA.html', 'r', encoding='utf-8') as f:
            return f.read()
    except FileNotFoundError:
        return "<h3 style='color:red;text-align:center;padding:20px;background:white;'>Erro: Arquivo 'CONSULTA.html' não encontrado na pasta!</h3>"


# --- SUAS ROTAS DE LOGICA DA API (MANTIDAS 100% INTACTAS) ---

@app.route('/api/consulta_direta', methods=['POST'])
def api_consulta_direta():
    d = request.json
    chave = d['chave']
    cache_file_path = os.path.join(CACHE_DIR, f"{chave}.json")
    if os.path.exists(cache_file_path):
        try:
            with open(cache_file_path, "r", encoding="utf-8") as f: cached_data = json.load(f)
            return jsonify({"success": True, "pdf": cached_data["pdf"], "xml": cached_data["xml"], "from_cache": True})
        except: pass
    try:
        r = requests.post(API_BASE_URL, json={'chave': chave}, headers={'Content-Type': 'application/json'}, timeout=15)
        if r.status_code != 200: return jsonify({"success": False, "reason": f"HTTP {r.status_code}"})
        res_data = r.json()
        pdf_base64 = res_data.get('pdf_base64', '')
        xml_bruto = res_data.get('xml', res_data.get('xml_base64', res_data.get('data', '')))
        if not pdf_base64 or not xml_bruto: return jsonify({"success": False, "reason": "Incompleto"})
        xml_str = str(xml_bruto).strip()
        xml_final = xml_str if xml_str.startswith('<') else base64.b64decode(xml_str).decode('utf-8')
        with open(cache_file_path, "w", encoding="utf-8") as f: json.dump({"pdf": pdf_base64, "xml": xml_final}, f)
        return jsonify({"success": True, "pdf": pdf_base64, "xml": xml_final, "from_cache": False})
    except: return jsonify({"success": False, "reason": "Erro"})

@app.route('/api/conciliar_duplo_restrito', methods=['POST'])
def api_conciliar_duplo_restrito():
    file_extrato = request.files['extrato']
    file_controle = request.files['controle']
    chaves_restritas = json.loads(request.form['chaves_lote'])
    try:
        import openpyxl
        wb_c = openpyxl.load_workbook(file_controle, data_only=True)
        sheet_c = wb_c.active
        headers_c = []
        header_c_row = 1
        for r_idx in range(1, 15):
            row_vals = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet_c[r_idx]]
            if 'data da viagem' in row_vals and 'prp' in row_vals and 'técnico' in row_vals:
                header_c_row = r_idx; headers_c = row_vals; break
        if not headers_c: return jsonify({"success": False, "reason": "Planilha de Controle incorreta."})
        idx_c_data = headers_c.index('data da viagem')
        idx_c_prp = headers_c.index('prp')
        idx_c_tecnico = headers_c.index('técnico')
        lista_pedidos_viagem = []
        for row in sheet_c.iter_rows(min_row=header_c_row + 1, values_only=True):
            if len(row) <= max(idx_c_data, idx_c_prp, idx_c_tecnico) or row[idx_c_data] is None or row[idx_c_tecnico] is None: continue
            lista_pedidos_viagem.append({
                "periodo_texto": str(row[idx_c_data]).strip().upper(),
                "tecnicos_texto": [t.strip() for t in str(row[idx_c_tecnico]).strip().upper().replace(' E ', ',').split(',')],
                "prp": str(row[idx_c_prp]).strip() if row[idx_c_prp] is not None else "-"
            })
        notas_lote = []
        for chave in chaves_restritas:
            file_path = os.path.join(CACHE_DIR, f"{chave}.json")
            if os.path.exists(file_path):
                try:
                    with open(file_path, "r", encoding="utf-8") as f: conteudo = json.load(f)
                    xml_texto = conteudo.get('xml', '')
                    valor_nota = 0.0; num_nota = "Não encontrado"; data_obj_nf = None; data_exibicao_nf = "-"
                    if '<vNF>' in xml_texto: valor_nota = float(xml_texto.split('<vNF>')[1].split('</vNF>')[0])
                    if '<nNF>' in xml_texto: num_nota = xml_texto.split('<nNF>')[1].split('</nNF>')[0]
                    if '<dhEmi>' in xml_texto:
                        dh_emi = xml_texto.split('<dhEmi>')[1].split('</dhEmi>')[0]
                        dt_clean = dh_emi.split('T')[0]; data_obj_nf = datetime.strptime(dt_clean, "%Y-%m-%d")
                        data_exibicao_nf = data_obj_nf.strftime("%d/%m/%Y")
                    notas_lote.append({"chave": chave, "valor": valor_nota, "numero": num_nota, "dt_obj": data_obj_nf, "data_completa": data_exibicao_nf})
                except: pass
        wb_e = openpyxl.load_workbook(file_extrato, data_only=True)
        sheet_e = wb_e.active
        headers_e = []
        header_e_row = 1
        for r_idx in range(1, 20):
            row_vals = [str(cell.value).strip().lower() if cell.value else "" for cell in sheet_e[r_idx]]
            if 'estabelecimento' in row_vals and 'saída' in row_vals: header_e_row = r_idx; headers_e = row_vals; break
        idx_e_portador = headers_e.index('nome do cartão')
        idx_e_valor = headers_e.index('saída')
        output_final = []
        for nota in notas_lote:
            valor_nota = nota["valor"]; numero_nf = nota["numero"]; data_obj_nf = nota["dt_obj"]; data_nota_pt = nota["data_completa"]; chave_nf = nota["chave"]
            nome_comprador = "Não localizado no Extrato"; prp_encontrada = "Não localizado no Controle"
            for row in sheet_e.iter_rows(min_row=header_e_row + 1, values_only=True):
                if len(row) <= max(idx_e_portador, idx_e_valor) or row[idx_e_valor] is None: continue
                val_extrato = abs(float(str(row[idx_e_valor]).replace(',', '.')))
                if abs(valor_nota - val_extrato) < 0.01:
                    nome_comprador = str(row[idx_e_portador]).strip(); comprador_clean = nome_comprador.upper().strip().split()[0]
                    if data_obj_nf:
                        for pedido in lista_pedidos_viagem:
                            if any(comprador_clean in t for t in pedido["tecnicos_texto"]):
                                try:
                                    if " A " in pedido["periodo_texto"]:
                                        p_inicio, p_fim = pedido["periodo_texto"].split(" A ")
                                        d_ini, m_ini = map(int, p_inicio.split("/")); d_fim, m_fim = map(int, p_fim.split("/"))
                                        dt_ini_viagem = datetime(data_obj_nf.year, m_ini, d_ini); dt_fim_viagem = datetime(data_obj_nf.year, m_fim, d_fim)
                                        if (dt_ini_viagem - timedelta(days=2)) <= data_obj_nf <= dt_fim_viagem: prp_encontrada = pedido["prp"]; break
                                    else:
                                        if data_obj_nf.strftime("%d/%m") in pedido["periodo_texto"]: prp_encontrada = pedido["prp"]; break
                                except: pass
                    break
            output_final.append({"chave": chave_nf, "valor_nota": valor_nota, "numero_nota": numero_nf, "data_nota_pt": data_nota_pt, "comprador": nome_comprador, "prp": prp_encontrada})
        return jsonify({"success": True, "results": output_final})
    except Exception as e:
        return jsonify({"success": False, "reason": f"Erro de Mapeamento: {str(e)}"})

if __name__ == '__main__':
    # Usa a porta fornecida pelo Railway ou a porta 8080 localmente
    port = int(os.environ.get("PORT", 8080))
    app.run(host='0.0.0.0', port=port, debug=True, threaded=True)